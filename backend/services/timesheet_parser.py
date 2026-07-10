"""Parser měsíčního timesheetu (.xlsx) — deterministický rozpad hodin podle kategorie.

Port validovaného prototypu (ověřeno proti reálným timesheetům/páskám 2026).
Očekávaný layout: sloupce A–E = DEN, DOBA, PŘESTÁVKY, STAV, DÉLKA; nový den má
ve sloupci A text ve tvaru "N. denname" (např. "3. pondělí"), pokračovací řádky
téhož dne mají sloupec A prázdný, řádek "Celkem…" ukončuje data.
"""
import calendar
import io
import re
from dataclasses import dataclass
from typing import Optional

from openpyxl import load_workbook

WEEKEND_DAYNAMES = ("sobota", "neděle")

# Noční pásmo 22:00–06:00 vyjádřené dvěma okny v dekadických hodinách od půlnoci.
# parse_range normalizuje rozsahy přes půlnoc přičtením 24 ke konci (20:00–02:00
# → [20, 26]), takže okno (22, 30) zachytí směnu běžící do noci a okno (-2, 6)
# směnu začínající až po půlnoci ([0, 4]). Nezjednodušovat na jedno okno.
NIGHT_WINDOWS = ((22.0, 30.0), (-2.0, 6.0))


@dataclass
class TimesheetHours:
    dov_h: float = 0.0            # dovolená
    prek_h: float = 0.0           # překážky / lékař
    volno_h: float = 0.0          # pracovní volno
    pres_wd: float = 0.0          # přesčas všední den
    pres_we: float = 0.0          # přesčas víkend
    svatek_h: float = 0.0         # práce ve svátek
    noc_h: float = 0.0            # noční hodiny v rámci přesčasu (22–06)
    pohot_h: float = 0.0          # pohotovost celkem
    pohot_overlap_h: float = 0.0  # překryv pohotovosti s přesčasem téhož dne
    worked_days: int = 0          # dny s odpracovanou směnou (pro stravenky)
    fond_days: int = 0            # Po–Pá dny nalezené v listu (diagnostika)
    total_hours: float = 0.0      # diagnostický součet, směna capnutá na 8 h/den


def parse_hours(value) -> float:
    """Délka záznamu: "8:00" → 8.0; "1 d"/"0,5 d" → násobky 8 h; jinak 0."""
    if value is None:
        return 0.0
    text = str(value).strip()
    m = re.search(r"(\d+):(\d+)", text)
    if m:
        return int(m.group(1)) + int(m.group(2)) / 60.0
    m = re.search(r"([\d,]+)\s*d", text)
    if m:
        return float(m.group(1).replace(",", ".")) * 8.0
    return 0.0


def parse_range(value) -> Optional[tuple[float, float]]:
    """Časový rozsah "HH:MM - HH:MM" → (start, end) v dekadických hodinách.

    Konec <= začátek znamená přechod přes půlnoc → end += 24.
    """
    if not value:
        return None
    m = re.search(r"(\d+):(\d+)\s*-\s*(\d+):(\d+)", str(value).strip())
    if not m:
        return None
    start = int(m.group(1)) + int(m.group(2)) / 60.0
    end = int(m.group(3)) + int(m.group(4)) / 60.0
    if end <= start:
        end += 24.0
    return (start, end)


def night_overlap(rng: Optional[tuple[float, float]]) -> float:
    """Hodiny rozsahu spadající do nočního pásma 22:00–06:00."""
    if rng is None:
        return 0.0
    start, end = rng
    total = 0.0
    for win_start, win_end in NIGHT_WINDOWS:
        lo = max(start, win_start)
        hi = min(end, win_end)
        if hi > lo:
            total += hi - lo
    return total


def range_overlap(r1: Optional[tuple[float, float]], r2: Optional[tuple[float, float]]) -> float:
    if not r1 or not r2:
        return 0.0
    return max(0.0, min(r1[1], r2[1]) - max(r1[0], r2[0]))


def compute_fond_days(year_month: str) -> int:
    """Počet Po–Pá dnů v měsíci ("YYYY-MM"). Svátky se nevyřazují — fond na
    páskách je počítá jako běžné pracovní dny (ověřeno 2025-01→23, 2025-02→20)."""
    year, month = (int(part) for part in year_month.split("-"))
    _, days_in_month = calendar.monthrange(year, month)
    return sum(
        1 for day in range(1, days_in_month + 1)
        if calendar.weekday(year, month, day) < 5
    )


def parse_timesheet(file_bytes: bytes) -> TimesheetHours:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.worksheets[0]

    # Seskupení řádků podle dne
    days: list[dict] = []
    current: Optional[dict] = None
    fond_days = 0
    for row in sheet.iter_rows(values_only=True):
        den = row[0] if len(row) > 0 else None
        if isinstance(den, str) and re.match(r"^\d+\.", den.strip()):
            parts = den.strip().split(".", 1)
            dayname = parts[1].strip() if len(parts) > 1 else ""
            current = {"dayname": dayname, "rows": [row]}
            days.append(current)
            if dayname not in WEEKEND_DAYNAMES:
                fond_days += 1
        elif den is not None and str(den).startswith("Celkem"):
            break
        elif current is not None and (den is None or den == ""):
            current["rows"].append(row)
    workbook.close()

    result = TimesheetHours(fond_days=fond_days)

    for day in days:
        weekend = day["dayname"] in WEEKEND_DAYNAMES
        pohot_ranges: list[tuple[float, float]] = []
        overtime_ranges: list[tuple[float, float]] = []
        day_work_h = 0.0
        worked = False
        day_other_h = 0.0

        for row in day["rows"]:
            doba = row[1] if len(row) > 1 else None
            stav = str(row[3]) if len(row) > 3 and row[3] is not None else ""
            delka = row[4] if len(row) > 4 else None
            hours = parse_hours(delka)
            rng = parse_range(doba)

            if "Pohotovost" in stav:
                result.pohot_h += hours
                day_other_h += hours
                if rng:
                    pohot_ranges.append(rng)
            elif "Přesčas" in stav:
                if weekend:
                    result.pres_we += hours
                else:
                    result.pres_wd += hours
                day_other_h += hours
                if rng:
                    overtime_ranges.append(rng)
                    result.noc_h += night_overlap(rng)
            elif "Dovolená" in stav:
                result.dov_h += hours
                day_other_h += hours
            elif "Svátek" in stav:
                result.svatek_h += hours
                day_other_h += hours
            elif "lékař" in stav or "Lékař" in stav or "Překážk" in stav:
                result.prek_h += hours
                day_other_h += hours
            elif ("Flexi" in stav or "8h" in stav
                    or "kolen" in stav or "eambuild" in stav):
                # Školení/Teambuilding matchované bez prvního písmene (Š/T versus
                # š/t v datech) — stejné substrings jako validovaný prototyp.
                day_work_h += hours
                worked = True
            elif "Pracovní volno" in stav:
                result.volno_h += hours
                day_other_h += hours

        if worked:
            result.worked_days += 1
        # Flexi + 8h záznamy popisují tutéž směnu — do diagnostického součtu
        # jde den max. jednou (8 h), ostatní kategorie bez capu.
        result.total_hours += min(8.0, day_work_h) + day_other_h

        for overtime in overtime_ranges:
            for pohot in pohot_ranges:
                result.pohot_overlap_h += range_overlap(overtime, pohot)

    return result
